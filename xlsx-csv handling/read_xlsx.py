## python script to clear csv files
import numpy as np
import openpyxl
import glob
import csv
inputnames = glob.glob('20230208_*.xlsx')
[k, l, m, n] = [1, 1, 1, 1]
for inputname in inputnames:
    print(inputname)
    wb = openpyxl.load_workbook(inputname)
    sh = wb.active # automatically selects Sheet1?
    i = 69
    j = 4
    SE1 = []
    SE2 = []
    velocity = sh.cell(row=6, column=3).value
    velocity = int(velocity.split()[0])
    while sh.cell(row=i, column=j).value != None:
        ## SENSOR ON/OFF
        if sh.cell(row=i, column=2).value == None and sh.cell(row=i+1, column=2).value != None:
            S1_on = sh.cell(row=i+1, column=1).value # SENSOR 1 ON
            # print(S1_on)
        if sh.cell(row=i, column=3).value != None and sh.cell(row=i+1, column=3).value == None:
            S2_off = sh.cell(row=i, column=1).value # SENSOR 2 OFF
            # print(S2_off)
        ##    
        ## TOTAL CPS for P1 and P2
        LE1 = sh.cell(row=i, column=j).value
        HE1 = sh.cell(row=i, column=j+1).value
        SE1.append((LE1 + HE1) * 10)
        LE2 = sh.cell(row=i, column=j+2).value
        HE2 = sh.cell(row=i, column=j+3).value
        SE2.append((LE2 + HE2) * 10)
        i+=1 # index for excel row
    SE1_cut = SE1[S1_on-1:S2_off]
    SE2_cut = SE2[S1_on-1:S2_off]
    # print(SE1)
    # print(SE2)

    ## Printed into single-column csv file
    ## Saved at current working directory
    ## SE1
    # outputname = 'MV_Co60_6' + str(k).zfill(2) + '_' + str(velocity) + '_P1.csv'
    # with open(outputname,'w+') as f:
    #     writer = csv.writer(f)
    #     csv.writer(f, delimiter='\n').writerow(SE1)
    #     k+=1
    ##
    ## SE1_cut
    # outputname = 'MV_Co60_6' + str(l).zfill(2) + '_' + str(velocity) + '_P1_cut.csv'
    # with open(outputname,'w+') as f:
    #     writer = csv.writer(f)
    #     csv.writer(f, delimiter='\n').writerow(SE1_cut)
    #     l+=1
    ##
    ## SE2
    # outputname = 'MV_Co60_6' + str(m).zfill(2) + '_' + str(velocity) + '_P2.csv'
    # with open(outputname,'w+') as f:
    #     writer = csv.writer(f)
    #     csv.writer(f, delimiter='\n').writerow(SE2)
    #     m+=1
    ##
    ## SE2_cut
    # outputname = 'MV_Co60_6' + str(n).zfill(2) + '_' + str(velocity) + '_P2_cut.csv'
    # with open(outputname,'w+') as f:
    #     writer = csv.writer(f)
    #     csv.writer(f, delimiter='\n').writerow(SE2_cut)
    #     n+=1
